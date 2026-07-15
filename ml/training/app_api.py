from __future__ import annotations

import copy
import csv
import os
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Any

import yaml
from data_manager.label_database import load_label_dataframe
from data_manager.metadata_stats import _build_line_sn_to_reference_map
from ml.dataset.label_filter import (
    LABEL_FILTER_STATUSES,
    _build_label_row_maps,
    _pick_training_label,
)
from data_manager.line_rules import LINE_RULES

from ml.train import (
    list_registered_model_types,
    list_supported_train_modes,
    model_catalog,
)

PROJECT_ROOT = Path(os.environ.get("FORVIA_REPO_ROOT", Path(__file__).resolve().parents[2])).expanduser()
CFG_DIR = Path(os.environ.get("FORVIA_CFG_DIR", PROJECT_ROOT / "cfg")).expanduser()
CORE_DATA_MANAGER_CONFIG = CFG_DIR / "core" / "data_manager.yaml"
CORE_LABEL_RULES_CONFIG = CFG_DIR / "core" / "label_rules.yaml"
SPLIT_STRATEGIES = ("reference_out", "line_out", "reference_in", "stratified")

DEFAULT_GROUPS = [
    ["正常", "干扰", "边界"],
    ["传感器错误"],
    ["秒表"],
    ["摩擦", "摩擦acc", "杂音"],
    ["咬齿"],
    ["震颤", "马达"],
    ["哒哒_咔咔"],
    ["其它"],
]
_DISTRIBUTION_CACHE: dict[tuple[str, float, str, float], dict] = {}
LINE_DISPLAY_ORDER = ("epump2", "epump3", "epump4", "etilt1", "etilt2")


def load_yaml(path: Path) -> dict:
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def data_manager_defaults() -> dict:
    cfg = load_yaml(CORE_DATA_MANAGER_CONFIG)
    data_root = str(cfg.get("data_root") or "")
    storage = str(cfg.get("tdms_storage_root") or "factory_raw")
    db = str(
        cfg.get("label_records_db_path")
        or cfg.get("sample_records_db_path")
        or ""
    )
    if "{data_root}" in db:
        db = db.replace("{data_root}", data_root.rstrip("/"))
    if not db and data_root:
        db = str(Path(data_root).expanduser() / "metadata" / "label_records.db")
    return {"data_root": data_root, "tdms_storage_root": storage, "label_records_db_path": db}


def reason_names() -> list[str]:
    cfg = load_yaml(CORE_LABEL_RULES_CONFIG)
    reasons = cfg.get("reasons") if isinstance(cfg.get("reasons"), dict) else {}
    return [str(v.get("name")) for v in reasons.values() if isinstance(v, dict) and v.get("name")]


def database_catalog(db_path_text: str, *, include_labels: bool = True) -> dict:
    db_path = Path(str(db_path_text or "")).expanduser()
    result = {"lines": [], "line_rows": [], "reason_rows": [], "status_rows": []}
    if not db_path.is_file():
        return result
    try:
        with sqlite3.connect(f"{db_path.resolve().as_uri()}?mode=ro", uri=True) as con:
            result["line_rows"] = [
                {"line": str(line), "samples": int(count)}
                for line, count in con.execute(
                    "SELECT line, COUNT(*) FROM samples WHERE is_active=1 GROUP BY line ORDER BY line"
                )
            ]
            result["lines"] = [row["line"] for row in result["line_rows"]]
            if include_labels:
                result["reason_rows"] = [
                    {"reason": str(reason or "未填写"), "count": int(count)}
                    for reason, count in con.execute(
                        """
                        SELECT COALESCE(NULLIF(e.reason_name, ''), NULLIF(e.reason_key, ''), '未填写'), COUNT(*)
                        FROM label_events e
                        WHERE e.status='confirmed'
                        GROUP BY 1 ORDER BY COUNT(*) DESC
                        """
                    )
                ]
                result["status_rows"] = [
                    {"status": str(status), "count": int(count)}
                    for status, count in con.execute(
                        "SELECT status, COUNT(*) FROM label_events GROUP BY status ORDER BY status"
                    )
                ]
    except Exception:
        return result
    return result


def _filtered_training_label_rows(db_path: Path) -> list[dict]:
    label_df = load_label_dataframe(db_path, statuses=LABEL_FILTER_STATUSES)
    if label_df.empty:
        return []
    grouped, fallback_grouped = _build_label_row_maps(label_df)
    if fallback_grouped:
        grouped = {
            **grouped,
            **{("", sn, sample_id): rows for (sn, sample_id), rows in fallback_grouped.items()},
        }
    out: list[dict] = []
    for (line, sn, sample_id), events in grouped.items():
        label, decision = _pick_training_label(events)
        if not label:
            continue
        event0 = events[0] if events else {}
        out.append(
            {
                "line": line or str(event0.get("line") or "").strip(),
                "sn": sn,
                "sample_id": sample_id,
                "decision": decision,
                **label,
            }
        )
    return out


def _count_rows(counter: Counter, key_name: str = "label") -> list[dict]:
    return [{key_name: str(name), "count": int(count)} for name, count in counter.most_common()]


def _label_distribution_text(rows: list[dict]) -> str:
    return "，".join(f"{item['label']}:{int(item['count'])}" for item in rows)


def _line_display_sort_key(line_name: object) -> tuple[int, str]:
    line = str(line_name or "").strip()
    try:
        return (LINE_DISPLAY_ORDER.index(line), line)
    except ValueError:
        return (len(LINE_DISPLAY_ORDER), line)


def _load_model_reference_map(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["明细(长表)"] if "明细(长表)" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        index = {name: headers.index(name) for name in ("产线", "机型", "reference", "模型策略") if name in headers}
        out: dict[tuple[str, str], dict[str, str]] = {}
        for row in rows:
            line = str(row[index["产线"]] or "").strip() if "产线" in index else ""
            reference = str(row[index["reference"]] or "").strip() if "reference" in index else ""
            if line and reference:
                out[(line, reference)] = {
                    "model_name": str(row[index["机型"]] or "").strip() if "机型" in index else "",
                    "strategy": str(row[index["模型策略"]] or "").strip() if "模型策略" in index else "",
                }
        workbook.close()
        return out
    except Exception:
        return {}


def _model_xlsx_path_from_manifest(manifest_path: Path) -> Path:
    candidate = manifest_path.parent / "型号数据.xlsx"
    if not candidate.is_file() and manifest_path.parent.name == "metadata":
        candidate = manifest_path.parent.parent / "型号数据.xlsx"
    return candidate


def database_distribution(db_path_text: str, manifest_path_text: str) -> dict:
    db_path = Path(str(db_path_text or "")).expanduser()
    manifest_path = Path(str(manifest_path_text or "")).expanduser()
    if not db_path.is_file() or not manifest_path.is_file():
        return {"line_rows": []}
    key = (
        str(db_path.resolve()),
        db_path.stat().st_mtime,
        str(manifest_path.resolve()),
        manifest_path.stat().st_mtime,
    )
    cached = _DISTRIBUTION_CACHE.get(key)
    if cached is not None:
        return copy.deepcopy(cached)
    with manifest_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        manifest_rows = list(csv.DictReader(f))
    labels = _filtered_training_label_rows(db_path)
    reference_by_sample, _ = _build_line_sn_to_reference_map(manifest_rows)
    model_map = _load_model_reference_map(_model_xlsx_path_from_manifest(manifest_path))
    line_map: dict[str, dict] = {}
    for row in manifest_rows:
        line = str(row.get("line") or "").strip()
        reference = str(row.get("reference") or "").strip() or "<EMPTY_reference>"
        entry = line_map.setdefault(line, {"line": line, "tdms_count": 0, "label_count": 0, "_labels": Counter(), "_labeled_sns": set(), "_refs": {}})
        entry["tdms_count"] += 1
        ref = entry["_refs"].setdefault(reference, {"reference": reference, "tdms_count": 0, "label_count": 0, "_labels": Counter(), "_labeled_sns": set()})
        ref["tdms_count"] += 1
    for row in labels:
        line = str(row.get("line") or "").strip()
        sn = str(row.get("sn") or "").strip()
        reference = reference_by_sample.get((line, sn), "<UNKNOWN_reference>")
        reason = str(row.get("reason_name") or row.get("reason_key") or row.get("result_name") or row.get("result_key") or "未标").strip()
        entry = line_map.setdefault(line, {"line": line, "tdms_count": 0, "label_count": 0, "_labels": Counter(), "_labeled_sns": set(), "_refs": {}})
        entry["label_count"] += 1
        entry["_labels"][reason] += 1
        if sn:
            entry["_labeled_sns"].add(sn)
        ref = entry["_refs"].setdefault(reference, {"reference": reference, "tdms_count": 0, "label_count": 0, "_labels": Counter(), "_labeled_sns": set()})
        ref["label_count"] += 1
        ref["_labels"][reason] += 1
        if sn:
            ref["_labeled_sns"].add(sn)
    line_rows = list(line_map.values())
    for line_row in line_rows:
        line_row["labeled_tdms_count"] = len(line_row.pop("_labeled_sns"))
        line_row["labels"] = _count_rows(line_row.pop("_labels"))
        line_row["label_distribution_text"] = _label_distribution_text(line_row["labels"])
        references = []
        model_groups: dict[str, dict] = {}
        for ref in line_row.pop("_refs").values():
            ref["labeled_tdms_count"] = len(ref.pop("_labeled_sns"))
            ref["labels"] = _count_rows(ref.pop("_labels"))
            ref["label_distribution_text"] = _label_distribution_text(ref["labels"])
            references.append(ref)
            mapping = model_map.get((line_row["line"], ref["reference"]), {})
            model_name = mapping.get("model_name") or "未配置模型"
            model = model_groups.setdefault(
                model_name,
                {"model_name": model_name, "strategy": mapping.get("strategy") or "", "tdms_count": 0, "label_count": 0, "labeled_tdms_count": 0, "_labels": Counter(), "references": []},
            )
            model["tdms_count"] += ref["tdms_count"]
            model["label_count"] += ref["label_count"]
            model["labeled_tdms_count"] += ref["labeled_tdms_count"]
            model["_labels"].update({item["label"]: item["count"] for item in ref["labels"]})
            model["references"].append(copy.deepcopy(ref))
        line_row["references"] = sorted(references, key=lambda x: (-int(x["tdms_count"]), x["reference"]))
        line_row["reference_count"] = len(line_row["references"])
        for model in model_groups.values():
            model["labels"] = _count_rows(model.pop("_labels"))
            model["label_distribution_text"] = _label_distribution_text(model["labels"])
            model["references"].sort(key=lambda x: (-int(x["tdms_count"]), x["reference"]))
        line_row["models"] = sorted(model_groups.values(), key=lambda x: (-int(x["tdms_count"]), x["model_name"]))
    line_rows.sort(key=lambda x: _line_display_sort_key(x.get("line")))
    total_labels = Counter()
    for line_row in line_rows:
        total_labels.update({item["label"]: item["count"] for item in line_row["labels"]})
    total = {
        "line_count": len(line_rows),
        "tdms_count": sum(int(row["tdms_count"]) for row in line_rows),
        "reference_count": sum(int(row["reference_count"]) for row in line_rows),
        "labeled_tdms_count": sum(int(row["labeled_tdms_count"]) for row in line_rows),
        "label_count": sum(int(row["label_count"]) for row in line_rows),
        "labels": _count_rows(total_labels),
    }
    total["label_distribution_text"] = _label_distribution_text(total["labels"])
    result = {
        "line_rows": line_rows,
        "total": total,
        "filter_rule": (
            "共用 ml/dataset/label_filter.py：expert 取最新；无 expert 时 operator 至少 2 条且标签完全一致才产生训练 y；"
            "operator 单标、operator 冲突、无人工标注不计入终标分布"
        ),
    }
    _DISTRIBUTION_CACHE.clear()
    _DISTRIBUTION_CACHE[key] = result
    return copy.deepcopy(result)


def line_reference_overview(config: dict) -> dict:
    normalize_database_input(config)
    database = config.get("database") if isinstance(config.get("database"), dict) else {}
    distribution = database_distribution(
        str(database.get("label_records_db_path") or ""),
        str(database.get("manifest_path") or ""),
    )
    line_rows = []
    for row in distribution.get("line_rows", []):
        line_rows.append(
            {
                key: copy.deepcopy(value)
                for key, value in row.items()
                if key != "models"
            }
        )
    return {
        "line_rows": line_rows,
        "filter_rule": distribution.get("filter_rule", ""),
    }


def xlsx_model_overview(config: dict) -> dict:
    normalize_database_input(config)
    database = config.get("database") if isinstance(config.get("database"), dict) else {}
    manifest_path = Path(str(database.get("manifest_path") or "")).expanduser()
    distribution = database_distribution(
        str(database.get("label_records_db_path") or ""),
        str(manifest_path),
    )
    return {
        "xlsx_path": str(_model_xlsx_path_from_manifest(manifest_path)),
        "line_rows": [
            {
                "line": row.get("line", ""),
                "models": copy.deepcopy(row.get("models", [])),
            }
            for row in distribution.get("line_rows", [])
        ],
        "filter_rule": distribution.get("filter_rule", ""),
    }


def _text_list(raw: object) -> list[str]:
    if isinstance(raw, str):
        return [raw.strip()] if raw.strip() else []
    return [str(item).strip() for item in (raw or []) if str(item).strip()]


def _key_sns_items(raw: object, *, default_line: str = "") -> list[dict[str, str]]:
    values = raw if isinstance(raw, list) else ([raw] if raw else [])
    out: list[dict[str, str]] = []
    for item in values:
        if isinstance(item, dict):
            path = str(item.get("path") or item.get("folder") or item.get("dir") or "").strip()
            line = str(item.get("line") or default_line).strip()
        else:
            path = str(item or "").strip()
            line = ""
        if path:
            out.append({"path": path, "line": line})
    return out


def training_data_options(config: dict) -> dict:
    normalize_database_input(config)
    database = config.get("database") if isinstance(config.get("database"), dict) else {}
    tdms_root = Path(str(database.get("tdms_root") or "")).expanduser()
    manifest_path = Path(str(database.get("manifest_path") or "")).expanduser()
    lines: set[str] = set()
    folders: set[str] = set()
    references: set[str] = set()
    if manifest_path.is_file():
        try:
            with manifest_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                for row in csv.DictReader(f):
                    line = str(row.get("line") or "").strip()
                    reference = str(row.get("reference") or "").strip()
                    relative_path = str(row.get("relative_path") or "").strip().replace("\\", "/")
                    if line:
                        lines.add(line)
                    if line and reference:
                        references.add(f"{line}/{reference}")
                    parent = str(Path(relative_path).parent).replace("\\", "/")
                    if parent and parent != ".":
                        folders.add(parent)
        except Exception:
            pass
    key_sns: set[str] = set()
    key_sns_by_line: dict[str, set[str]] = {}
    current_line = str(config.get("line_name") or "").strip()
    def infer_key_line(path: Path, fallback: str = "") -> str:
        try:
            parts = path.resolve().relative_to(tdms_root.resolve()).parts
        except (OSError, ValueError):
            return fallback
        if not parts:
            return fallback
        if parts[0] == "prototype" and len(parts) > 1:
            return parts[1]
        if parts[0]:
            return parts[0]
        return fallback

    def add_key_path(path: Path, path_line: str) -> None:
        if not path.is_dir():
            return
        text = str(path)
        key_sns.add(text)
        if path_line:
            key_sns_by_line.setdefault(path_line, set()).add(text)

    prototype_root = tdms_root / "prototype"
    if prototype_root.is_dir():
        add_key_path(prototype_root, current_line)
        try:
            for path in prototype_root.iterdir():
                if path.is_dir():
                    add_key_path(path, path.name if path.name in lines else current_line)
        except OSError:
            pass
    for line in lines | {current_line}:
        root = tdms_root / line / "key_sns"
        if root.is_dir():
            add_key_path(root, line)
            try:
                for path in root.rglob("*"):
                    add_key_path(path, line)
            except OSError:
                pass
    data = config.get("data") if isinstance(config.get("data"), dict) else {}
    folders.update(_text_list(data.get("folders") or data.get("folder")))
    selected_key_sns: list[dict[str, str]] = []
    selected_key_line = str(data.get("key_line") or current_line).strip()
    for item in _key_sns_items(data.get("key_sns"), default_line=selected_key_line):
        path = Path(item["path"]).expanduser()
        item_line = item["line"] or infer_key_line(path, selected_key_line)
        add_key_path(path, item_line)
        selected_key_sns.append({"path": str(path), "line": item_line})
    lines.update(key_sns_by_line)
    key_sns_items = [
        {"path": path, "line": line}
        for line, paths in sorted(key_sns_by_line.items())
        for path in sorted(paths)
    ]
    return {
        "lines": sorted(x for x in lines if x),
        "folders": sorted(x for x in folders if x),
        "references": sorted(x for x in references if x),
        "key_sns": sorted(x for x in key_sns if x),
        "key_sns_by_line": {
            line: sorted(paths)
            for line, paths in sorted(key_sns_by_line.items())
            if line
        },
        "key_sns_items": key_sns_items,
        "selected_key_sns": selected_key_sns,
    }


def training_app_options() -> dict:
    defaults = data_manager_defaults()
    models = model_catalog()
    return {
        **models,
        "split_strategies": list(SPLIT_STRATEGIES),
        "reasons": reason_names(),
        "line_options": sorted(LINE_RULES),
        "database_defaults": defaults,
        "database_catalog": database_catalog(
            defaults.get("label_records_db_path", ""),
            include_labels=False,
        ),
    }


def default_config(line_name: str, model_name: str, model_type: str) -> dict:
    defaults = data_manager_defaults()
    default_reasons = reason_names() or [reason for group in DEFAULT_GROUPS for reason in group]
    return {
        "line_name": line_name,
        "model": {"model_name": model_name},
        "results_path": "./results/",
        "database": {
            "label_records_db_path": defaults["label_records_db_path"],
            "tdms_root": str(Path(defaults["data_root"]).expanduser() / defaults["tdms_storage_root"])
            if defaults["data_root"] else "",
            "manifest_path": str(Path(defaults["data_root"]).expanduser() / "metadata" / "tdms_manifest.csv")
            if defaults["data_root"] else "",
            "configured": False,
            "last_operation": "",
        },
        "data_root": defaults["data_root"],
        "label_records_db_path": defaults["label_records_db_path"],
        "data": {"line": [line_name], "folders": [], "reference": [], "key_sns": []},
        "dataset": {
            "batch_size": 2000,
            "num_workers": 8,
            "data_root": defaults["data_root"],
            "label_records_db_path": defaults["label_records_db_path"],
        },
        "train": {
            "model_type": model_type,
            "train_mode": "normal",
            "split_strategy": "reference_in",
            "seed_runs": 10,
            "label_mapping": {
                "enable": True,
                "auto_mlabel": True,
                "extract_all_features": False,
                "groups": [
                    {"weight": 1.0, "reasons": [{"reason": reason, "sample": -1}]}
                    for reason in default_reasons
                ],
            },
            "class_weights": {reason: 1.0 for reason in default_reasons},
        },
    }


def normalize_project_payload(data: dict) -> dict:
    line_name = str(data.get("line_name") or "").strip()
    model_name = str(data.get("model_name") or "").strip()
    model_type = str(data.get("model_type") or "xgb").strip().lower()
    if not line_name or not model_name:
        raise ValueError("line_name 和 model_name 不能为空")
    if model_type not in list_registered_model_types():
        raise ValueError(f"不支持的 model_type: {model_type}")
    config = copy.deepcopy(data.get("config") if isinstance(data.get("config"), dict) else {})
    if not config:
        config = default_config(line_name, model_name, model_type)
    config["line_name"] = line_name
    config.setdefault("model", {})["model_name"] = model_name
    train = config.setdefault("train", {})
    train["model_type"] = model_type
    train.setdefault("train_mode", "normal")
    train.setdefault("split_strategy", "reference_in")
    train.setdefault("seed_runs", 10)
    config.setdefault("results_path", "./results/")
    normalize_database_input(config)
    name = f"{line_name}_{model_name}"
    return {
        "name": name,
        "line_name": line_name,
        "model_name": model_name,
        "model_type": model_type,
        "config": config,
    }


def normalize_database_input(config: dict) -> dict:
    """以新的 label_records.db 为主输入，并补齐现有训练脚本需要的兼容字段。"""
    defaults = data_manager_defaults()
    database = config.setdefault("database", {})
    db_text = str(
        database.get("label_records_db_path")
        or config.get("label_records_db_path")
        or (config.get("dataset") or {}).get("label_records_db_path")
        or database.get("sample_records_db_path")
        or config.get("sample_records_db_path")
        or (config.get("dataset") or {}).get("sample_records_db_path")
        or defaults.get("label_records_db_path")
        or ""
    ).strip()
    if db_text and Path(db_text).name == "sample_records.db":
        db_text = str(Path(db_text).with_name("label_records.db"))
    tdms_root_text = str(
        database.get("tdms_root")
        or (
            Path(str(defaults.get("data_root") or "")).expanduser()
            / str(defaults.get("tdms_storage_root") or "factory_raw")
            if defaults.get("data_root")
            else ""
        )
    ).strip()
    manifest_text = str(database.get("manifest_path") or "").strip()
    data_root_text = ""

    db_path = Path(db_text).expanduser() if db_text else None
    tdms_root = Path(tdms_root_text).expanduser() if tdms_root_text else None
    if tdms_root is not None:
        data_root_text = str(tdms_root.parent)
    elif db_path is not None and db_path.parent.name == "metadata":
        data_root_text = str(db_path.parent.parent)
    else:
        data_root_text = str(config.get("data_root") or "").strip()
    if not manifest_text and db_path is not None:
        manifest_text = str(db_path.parent / "tdms_manifest.csv")

    database["label_records_db_path"] = db_text
    database["tdms_root"] = tdms_root_text
    database["manifest_path"] = manifest_text
    config["label_records_db_path"] = db_text
    config["data_root"] = data_root_text
    dataset = config.setdefault("dataset", {})
    dataset["label_records_db_path"] = db_text
    dataset["data_root"] = data_root_text
    dataset["manifest_path"] = manifest_text
    return config


def model_id(config: dict) -> str:
    line = str(config.get("line_name") or "").strip()
    model = config.get("model") if isinstance(config.get("model"), dict) else {}
    model_name = str(model.get("model_name") or "").strip()
    return f"{line}_{model_name}".strip("_")


def validate_config(config: dict) -> dict:
    errors: list[str] = []
    warnings: list[str] = []
    train = config.get("train") if isinstance(config.get("train"), dict) else {}
    if not str(config.get("line_name") or "").strip():
        errors.append("缺少 line_name")
    if not str((config.get("model") or {}).get("model_name") or "").strip():
        errors.append("缺少 model.model_name")
    if train.get("model_type") not in list_registered_model_types():
        errors.append("train.model_type 无效")
    if train.get("train_mode") not in list_supported_train_modes():
        errors.append("train.train_mode 无效")
    if train.get("split_strategy") not in SPLIT_STRATEGIES:
        errors.append("train.split_strategy 无效")
    groups = ((train.get("label_mapping") or {}).get("groups") or [])
    if not groups:
        errors.append("train.label_mapping.groups 不能为空")
    for index, group in enumerate(groups, start=1):
        try:
            weight = float(group.get("weight", 1.0))
        except (TypeError, ValueError):
            errors.append(f"标签组 {index} 的 weight 必须是数字")
            continue
        if weight <= 0:
            errors.append(f"标签组 {index} 的 weight 必须大于 0")
    class_weights = train.get("class_weights")
    if class_weights is not None and not isinstance(class_weights, dict):
        errors.append("train.class_weights 必须是类别名称到权重的映射")
    elif isinstance(class_weights, dict):
        for name, raw_weight in class_weights.items():
            try:
                weight = float(raw_weight)
            except (TypeError, ValueError):
                errors.append(f"类别 {name} 的权重必须是数字")
                continue
            if weight <= 0:
                errors.append(f"类别 {name} 的权重必须大于 0")
    early_stopping_metric = str(train.get("early_stopping_metric") or "").strip()
    early_stopping_rounds = train.get("early_stopping_rounds")
    if early_stopping_metric and early_stopping_metric not in {"f1_macro", "eval_metric"}:
        errors.append("train.early_stopping_metric 仅支持 f1_macro 或 eval_metric")
    if early_stopping_rounds is not None:
        try:
            if int(early_stopping_rounds) <= 0:
                errors.append("train.early_stopping_rounds 必须大于 0；关闭早停请删除该参数")
        except (TypeError, ValueError):
            errors.append("train.early_stopping_rounds 必须是整数")
    data_cfg = config.get("data") if isinstance(config.get("data"), dict) else {}
    if not any(_text_list(data_cfg.get(key)) for key in ("line", "folders", "folder", "reference")):
        errors.append("训练主数据必须选择 line、folders 或 reference 之一")
    normalize_database_input(config)
    database = config.get("database") if isinstance(config.get("database"), dict) else {}
    db_path = Path(str(database.get("label_records_db_path") or "")).expanduser()
    tdms_root = Path(str(database.get("tdms_root") or "")).expanduser()
    manifest_path = Path(str(database.get("manifest_path") or "")).expanduser()
    data_root = Path(str(config.get("data_root") or "")).expanduser()
    if not db_path.is_file():
        errors.append(f"新的 label_records.db 不存在: {db_path}")
    if not tdms_root.is_dir():
        warnings.append(f"TDMS 根目录不存在: {tdms_root}")
    if not manifest_path.is_file():
        errors.append(f"tdms_manifest.csv 不存在: {manifest_path}")
    if not data_root.exists():
        warnings.append(f"无法从数据库或 TDMS 根目录推断有效 data_root: {data_root}")
    data_cfg = config.get("data") or {}
    raw_key_items = data_cfg.get("key_sns") if isinstance(data_cfg.get("key_sns"), list) else []
    key_items = _key_sns_items(data_cfg.get("key_sns"), default_line=str(data_cfg.get("key_line") or ""))
    for raw in _text_list(data_cfg.get("folders") or data_cfg.get("folder")) + [item["path"] for item in key_items]:
        if raw and not Path(str(raw)).expanduser().exists():
            warnings.append(f"数据路径不存在: {raw}")
    for index, item in enumerate(key_items):
        path = Path(item["path"]).expanduser()
        if path.exists() and not path.is_dir():
            errors.append(f"data.key_sns 必须选择文件夹，不能选择文件: {path}")
        inferred_line = ""
        if tdms_root.is_dir():
            try:
                parts = path.resolve().relative_to(tdms_root.resolve()).parts
                if parts:
                    inferred_line = parts[1] if parts[0] == "prototype" and len(parts) > 1 else parts[0]
            except (OSError, ValueError):
                pass
        raw_item = raw_key_items[index] if index < len(raw_key_items) else None
        if isinstance(raw_item, dict) and not str(raw_item.get("line") or "").strip():
            errors.append(f"key 数据必须选择 line: {path}")
        elif not item["line"] and not inferred_line:
            errors.append(f"key 数据必须选择 line: {path}")
    return {"ok": not errors, "errors": errors, "warnings": warnings, "model_id": model_id(config)}


def dump_yaml(config: dict) -> str:
    return yaml.safe_dump(config, allow_unicode=True, sort_keys=False)


def config_filename(line_name: str, model_name: str) -> str:
    safe = "_".join(
        part for part in "".join(
            char if char.isalnum() or char in ("_", "-") else "_"
            for char in f"{line_name}_{model_name}"
        ).split("_") if part
    )
    return f"{safe or 'model'}.yaml"


def config_path_for_project(payload: dict, existing_path: str = "") -> Path:
    if existing_path:
        return Path(existing_path).expanduser()
    return CFG_DIR / config_filename(payload["line_name"], payload["model_name"])


def save_project_config(payload: dict, existing_path: str = "") -> dict:
    """Directly persist the project configuration to its cfg YAML file."""
    payload = copy.deepcopy(payload)
    payload["config"]["results_path"] = "./results/"

    path = config_path_for_project(payload, existing_path)
    if not existing_path and path.exists():
        raise ValueError(f"配置 YAML 已存在: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(dump_yaml(payload["config"]), encoding="utf-8")
    payload["config_path"] = str(path.resolve())
    return payload


def parse_yaml(text: str) -> dict:
    data = yaml.safe_load(text) or {}
    if not isinstance(data, dict):
        raise ValueError("YAML 顶层必须是对象")
    return data


def dataset_summary(config: dict) -> dict:
    normalize_database_input(config)
    database = config.get("database") if isinstance(config.get("database"), dict) else {}
    data_root = Path(str(config.get("data_root") or "")).expanduser()
    tdms_root = Path(str(database.get("tdms_root") or "")).expanduser()
    db = Path(str(database.get("label_records_db_path") or "")).expanduser()
    manifest = Path(str(database.get("manifest_path") or data_root / "metadata" / "tdms_manifest.csv")).expanduser()
    data_paths = []
    data_cfg = config.get("data") if isinstance(config.get("data"), dict) else {}
    for key in ("folders", "key_sns"):
        values = data_cfg.get(key) or (data_cfg.get("folder") if key == "folders" else [])
        raw_values = _text_list(values) if key == "folders" else [item["path"] for item in _key_sns_items(values)]
        for raw in raw_values:
            p = Path(str(raw)).expanduser()
            data_paths.append({"path": str(p), "exists": p.exists(), "kind": key})
    manifest_rows = 0
    if manifest.exists():
        try:
            with manifest.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                manifest_rows = sum(1 for _ in csv.DictReader(f))
        except Exception:
            pass
    sample_rows = label_rows = confirmed_labels = 0
    if db.is_file():
        try:
            with sqlite3.connect(db) as con:
                sample_rows = int(con.execute("SELECT COUNT(*) FROM samples WHERE is_active=1").fetchone()[0])
                label_rows = int(con.execute("SELECT COUNT(*) FROM label_events").fetchone()[0])
                confirmed_labels = int(con.execute("SELECT COUNT(*) FROM label_events WHERE status='confirmed'").fetchone()[0])
        except Exception:
            pass
    catalog = database_catalog(str(db))
    return {
        "data_root": str(data_root),
        "data_root_exists": data_root.exists(),
        "tdms_root": str(tdms_root),
        "tdms_root_exists": tdms_root.exists(),
        "label_records_db_path": str(db),
        "label_records_db_exists": db.exists(),
        "sample_rows": sample_rows,
        "label_rows": label_rows,
        "confirmed_labels": confirmed_labels,
        "manifest_path": str(manifest),
        "manifest_exists": manifest.exists(),
        "manifest_rows": manifest_rows,
        "paths": data_paths,
        **catalog,
    }
